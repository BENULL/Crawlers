import requests
from pyquery import PyQuery as pq
import pandas as pd
import time
import sys


info = pd.DataFrame({'链接': [],'标题': [],'中标商机构': [], '中标商地址': [], '中标金额': [], '包名': [], '项目名称': [], '中标总价': [], '采购人': [],'联系地址':[]})

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def add_data(index, url, title, zb_name, zb_address, zb_money, b_name,p_name,zb_sum,cg_person,cg_address):
    global info
    info.loc[index] = [url, title, zb_name, zb_address, zb_money, b_name,p_name,zb_sum,cg_person,cg_address]

def handle_resp2(resp,index,url,title):
    # 处理表格信息
    resp_data = resp.text
    doc = pq(resp_data)
    tables = doc('.wenben table')
    
    for table in tables.items():
        print(table)
        tb_pd = pd.read_html(str(table),encoding='utf-8')
        print(tb_pd)

def handle_resp(resp,index,url,title):
    resp_data = resp.text
    doc = pq(resp_data)
    info_div = doc('#tab3 .Article')
    #print(info_div)
    zb_name = ''
    zb_num = 0
    zb_address = ''
    zb_address_num = 0
    cg_person = ''
    is_cgperson_get = False
    cg_address = ''
    zb_money = ''
    # zb_money_num=''
    b_name=''
    p_name=''
    zb_sum=''
    is_zb_address =False
    # 处理行信息
    #rows = info_div.children('p').items()
    # if l == 0:
    rows = doc('#tab3 .Article div').children('p').items()

    for i in rows:
        ss=i.text()
        strs= ss.split('\n')
        for s in strs:
            s=s.replace(' ','').replace('\n', '')
            print(s)
            if '中标供应商名称' in s or '成交供应商名称' in s or '全称' in s or '成交供应商：' in s or '中标单位' in s:
                pos = s.find('：')
                if zb_num>0:
                    zb_name =zb_name+'*'+s[pos+1:]
                    zb_num += 1
                else:
                    zb_name+=s[pos+1:]
                    zb_num+=1
                is_zb_address = True
                continue
            if '中标供应商地址' in s or '成交供应商地址' in s or '地址' in s :
                pos = s.find('：')
                if is_zb_address:
                    if zb_num>1 :
                        zb_address = zb_address + '*' + s[pos+1:]
                    elif zb_num==1:
                        zb_address= s[pos+1:]
                is_zb_address = False
                continue
            if '中标金额' in s or '成交金额' in s or '中标总额' in s or '金额' in s or  '总额' in s:
                start,end=-1,-1
                for i,c in enumerate(s):
                    if c.isdigit() and start == -1:
                        start=i
                    if start>=0 and (c=='.' or c.isdigit()) :
                        end=i
                    if start >= 0 and not (c == '.' or c.isdigit()):
                        break
                if zb_money :
                    zb_money = zb_money + '*' + s[pos + 1:]
                else:
                    zb_money = s[start:end+1]
                continue
            if '采购人' in s:
                is_cgperson_get = True
                pos = s.find('：')
                cg_person = s[pos+1:]
                continue

            if '联系地址' in s and is_cgperson_get:
                pos = s.find('：')
                cg_address = s[pos+1:]
                continue
    # 处理表格信息
    tables = doc('#tab3 table')
    for table in tables.items():
        tb_pd = pd.read_html(str(table))
        for t in tb_pd:
            print(t)
            col_list = list(t.loc[0].values)
            p_name_index = -1
            b_name_index = -1
            zb_sum_index = -1
            for i,col in enumerate(col_list):
                if '名称' in col:
                    p_name_index = i
                if '序号' in col:
                    b_name_index = i
                if '价' in col or '额' in col:
                    zb_sum_index = i
            if p_name_index>=0:
                p_name_list = t[p_name_index][1:].tolist()
                if p_name:
                    p_name+='*'
                p_name += '*'.join(p_name_list)
            if b_name_index>=0:
                b_name_list = t[b_name_index][1:].tolist()
                if b_name:
                    b_name+='*'
                b_name +='*'.join(b_name_list)
            if zb_sum_index>=0:
                zb_sum_list = t[zb_sum_index][1:].tolist()
                if zb_sum:
                    zb_sum+='*'
                zb_sum += '*'.join(zb_sum_list)


    print('========================'+zb_name)
    print('========================' + zb_address)
    print('========================' + zb_money)
    print('========================' + cg_person)
    print('========================' + cg_address)
    print('========================' + b_name)
    print('========================' + p_name)
    print('========================' + zb_sum)
    add_data(index, url, title, zb_name, zb_address, zb_money, b_name,p_name,zb_sum,cg_person,cg_address)




headers = {
    # 常规操作，可以套用
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
}
sess = requests.session()
#datas = pd.read_excel(r'/Users/null/Downloads/load.xlsx', sheet_name = '仅一个表格')
#total_lenght = datas.shape[0]
#io ='./test10.xlsx'
#append_df_to_excel(io, info, header=True)
#for index,row in datas.iterrows():
for index in range(1):
    try:
        url = 'http://www.ccgp-hebei.gov.cn/bd/bd_by/cggg/zhbggAAAA/202006/t20200608_1236479.html'
        
        title= 'test'
        #url = row['链接']
        # title= row['标题']
        # print(f'爬去{url}中标文件{title}')
        resp = sess.get(url, timeout=(2, 9), headers=headers)
        #handle_resp(resp,index,url,title)
        handle_resp2(resp,index,url,title)

        # if index % 100 == 0 or index == total_lenght-1:
        #     print("保存一次", index)
        #     append_df_to_excel(io, info, header=False)
        #     info = pd.DataFrame({'链接': [], '标题': [], '中标商机构': [], '中标商地址': [], '中标金额': [], '包名': [], '项目名称': [], '中标总价': [], '采购人': [],'联系地址': []})

    except Exception as e:
        print(e)
        print(f'爬去{url}中标文件{title}出现异常')

