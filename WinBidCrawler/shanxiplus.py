# shanxizb.py
import requests
from pyquery import PyQuery as pq
import pandas as pd
import time
import sys


info = pd.DataFrame({'页码': [], '当前页位置': [], '地区': [], '类型': [], '名称': [], '发布时间': [], '链接': []})

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def add_data(index, n, area, style, name, chengjiaodate, href_url):
    global info
    id = str(index) + '-' + str(n + 1)
    info.loc[id] = [index, n + 1, area, style, name, chengjiaodate, href_url]


def handle_resp(resp, index):
    resp_data = resp.text
    # print(resp_data)
    doc = pq(resp_data)
    lis = list(doc('#node_list tbody tr').items())
    # print(lis)
    for n in range(0, len(lis)):
        area = lis[n]('td:nth-child(2)').text()
        style = '中标公告'
        name = lis[n]('td:nth-child(1)').text()
        href = str(lis[n]('td:nth-child(1) a').attr('href'))
        href_url = href.replace('view.php?nid=','http://www.ccgp-shanxi.gov.cn/view.php?nid=')
        chengjiaodate = lis[n]('td:nth-child(4)').text()
        # print(index,n,area,style,name,chengjiaodate,href_url)
        add_data(index,n,area,style,name,chengjiaodate,href_url)
        n = n+1

headers = {
    # 常规操作，可以套用
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
}
sess = requests.session()
query_url1 = "http://www.ccgp-shanxi.gov.cn/view.php?app=&type=&nav=104&page="

begin_index = int(sys.argv[1])
total_lenght = int(sys.argv[2])
index = int(sys.argv[1])
while index<total_lenght:
    try:
        index = index + 1
        url = query_url1 + str(index)
        # print(url)
        print('--爬取第', index, '页')
        resp = sess.get(url, timeout=(2, 9), headers=headers)
        # print(resp)
        handle_resp(resp, index)
        time.sleep(0.4)
        if index % 100 == 0 or index == total_lenght:
            print("保存一次", index)
            append_df_to_excel('test3.xlsx',info,header=False)
            info = pd.DataFrame({'页码': [], '当前页位置': [], '地区': [], '类型': [], '名称': [], '发布时间': [], '链接': []})

            #info.to_excel('test1.xlsx')
        index = index + 1
    except Exception as e:
        print(e)
        print('--爬取第', index, '页出现异常')
    finally:
        index=index-1

